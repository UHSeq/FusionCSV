# sort and merge multiple csv files to one single file for import into excel.

import os
from zipfile import ZipFile
import pandas
import openpyxl
from datetime import datetime

cwd = os.getcwd()
now = datetime.now()
nowtime = now.strftime('%H%M%S')
nowdate = now.strftime('%m%d%Y')
excel_file_name = f'fusions_{nowtime}_{nowdate}.xlsx'
output_dir = 'Output'
csv_hold = 'csv_hold'
excel_path = os.path.join(cwd, output_dir, excel_file_name)
count_col = 3

zip_files = []
csv_files = set()
csv_objs = {}


class CSVFiles():
    def __init__(self, path):
        self.path = path
        _, self.name = os.path.split(self.path)
        self.name, _ = os.path.splitext(self.name)
        self.read_data()
        self.find_groups()
        self.process_groups()
        self.data.to_csv(os.path.join(cwd, csv_hold, self.name + '.csv'))

    def read_data(self):
        self.data = pandas.read_csv(self.path)

    def find_groups(self):
        self.groupby = self.data.groupby(['Name'])
        self.size = self.groupby.size()
        self.size = dict(self.size)
        self.data.insert(loc=count_col, column='Counts', value=0)

    
    def process_groups(self):
        for index, name in enumerate(self.data['Name']):
            self.data.iat[index, count_col] = self.size[name]
            # print(index, name)

# class Uniques():
#     def __init__(self, name) -> None:
#         self.name = name

class MasterData():
    def __init__(self, path):
        self.path = path
        self.create_excel()
        self.csv_files = {}
        self.fusions = {}
        self.csv_dict()
        self.fusion_count()

    def csv_dict(self):
        for index, csv in enumerate(csv_files, 1):
            print(f'####\nInitializing {index}\n{csv}\n####')
            self.csv_files[index] = CSVFiles(csv)
            # print()

    def create_excel(self):
        workbook = openpyxl.Workbook()
        workbook.save(self.path)
        self.workbook = openpyxl.load_workbook(self.path)

    def remove_sheet(self):
        # self.workbook = openpyxl.load_workbook(self.path)
        self.workbook.remove(self.workbook['Sheet'])
        self.workbook.save(self.path)

    def rename_sheet(self):
        wso = self.workbook['Sheet']
        wso.title('Total Fusions')
        # pass

    def fusion_count(self):
        # pass
        # csv_headers = ['Name']
        # self.fusions = []
        for index, csv in self.csv_files.items():
            # print(index, csv)
            # print(index, type(csv.size))
            # print(index, csv.name)
            for fusion, count in csv.size.items():
                pass
                if not fusion in self.fusions:
                    self.fusions[fusion] = {}
                # # else:
                # #     self.fusions[fusion] = {}
                if count > 0:
                    self.fusions[fusion][csv.name] = count
                else:
                    self.fusions[fusion][csv.name] = 0
                # print(fusion, count)
                # self.fusions[fusion] = {}
                # self.fusions
                # print(fusion, count)
                # csv_headers.append(csv.name)
                # self.fusions
        print(self.fusions)
        # self.fusions = pandas.DataFrame.from_dict(self.fusions)
        # self.fusions.to_csv(os.path.join(cwd, output_dir, 'temp_merged.csv'))


def add_zip(directory, name, ziplist=zip_files):
    zip_path = write_pathname(directory, name)
    ziplist.append(zip_path)

def add_csv(directory, name, csvset=csv_files):
    csv_path = write_pathname(directory, name)
    csvset.add(csv_path)

def clear_folder(dir_path):
    pass
    for root, _, files in os.walk(dir_path):
        for file in files:
            os.remove(os.path.join(root, file))

# def initialize_csv():
#     # workbook = openpyxl.Workbook()
#     # workbook.save(excel_path)
#     for index, csv in enumerate(csv_files, 1):
#         print(f'#### Initializing {index} ####')
#         csv_objs[index] = CSVFiles(csv)
#     # workbook = openpyxl.load_workbook(excel_path)
#     # workbook.remove(workbook['Sheet'])
#     # workbook.save(excel_path)

def unzipper(zipperlist, csvset):
    for zipperfile in zipperlist:
        zipperpath = os.path.dirname(zipperfile)
        with ZipFile(zipperfile, 'r') as Zipper:
            tempset = set(ZipFile.namelist(Zipper))
            for file in tempset:
                csvset.add(os.path.join(zipperpath, file))
            Zipper.extractall(os.path.dirname(zipperfile))

def write_pathname(*strings):
    return os.path.join(*strings)

def main_loop():
    clear_folder(output_dir)
    clear_folder(csv_hold)
    for dir, _, file in os.walk(cwd):
        for name in file:
            if name.endswith('.csv'):
                add_csv(dir, name)
            elif name.endswith('.zip'):
                add_zip(dir, name)
    if len(zip_files) > 0:
        unzipper(zip_files, csv_files)
        # pass
    if len(csv_files) > 0:
        # initialize_csv()
        m = MasterData(excel_path)
        # pass
    else:
        print("####\nOpps... did you mean to add some csv files?\nNothing here...\n####")

if __name__ in '__main__':
    main_loop()
