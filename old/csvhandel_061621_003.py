# sort and merge multiple csv files to one single file for import into excel.

import os
from zipfile import ZipFile
import pandas
import openpyxl
# import zipfile

cwd = os.getcwd()
excel_path = os.path.join(cwd, 'Samples', 'fusions.xlsx')
# print(cwd, type(cwd))

zip_files = []
csv_files = set()
csv_objs = {}


class CSVFiles():
    def __init__(self, path):
        # pass
        self.path = path
        _, self.name = os.path.split(self.path)
        self.name, _ = os.path.splitext(self.name)
        self.read_data()
        self.find_unique()

    def read_data(self):
        self.data = pandas.read_csv(self.path)
        # print(type(self.data))

    def find_unique(self):
        # self.unique_names = self.data.Name.unique()
        self.unique_counts = self.data.groupby(['Name']).size()
        # self.unique_counts.to_csv(os.path.join(cwd, 'Samples', self.name))
        # self.data.to_excel(writer, sheet_name=self.name)
        with pandas.ExcelWriter(excel_path, mode='a') as Excel_File:
            # Excel_File.book = openpyxl.load_workbook(excel_path)
            self.unique_counts.to_excel(Excel_File, sheet_name=self.name)
        # with openpyxl.Workbook()
        # workbook.
        # workbook.save(excel_path)



class Uniques():
    def __init__(self, name) -> None:
        # pass
        self.name = name


def write_pathname(*strings):
    return os.path.join(*strings)

def add_zip(directory, name, ziplist=zip_files):
    zip_path = write_pathname(directory, name)
    ziplist.append(zip_path)
    # pass

def add_csv(directory, name, csvset=csv_files):
    csv_path = write_pathname(directory, name)
    csvset.add(csv_path)
    # pass

def unzipper(zipperlist, csvset):
    for zipperfile in zipperlist:
        zipperpath = os.path.dirname(zipperfile)
        with ZipFile(zipperfile, 'r') as Zipper:
            tempset = set(ZipFile.namelist(Zipper))
            for file in tempset:
                csvset.add(os.path.join(zipperpath, file))
            Zipper.extractall(os.path.dirname(zipperfile))

def initialize_csv():
    # pass
    # with pandas.ExcelWriter(excel_path) as Excel_File:
    #     Excel_File.book = openpyxl.load_workbook(excel_path)
    workbook = openpyxl.Workbook()
    workbook.save(excel_path)
    workbook.close()
    for index, csv in enumerate(csv_files):
        csv_objs[index] = CSVFiles(csv)
        # print(index, csv)
    # workbook = openpyxl.Workbook()
    del workbook['Sheet']
    # del first_sheet
    print(workbook.sheetnames)
    workbook.save(excel_path)


def test_loop():
    for dir, _, file in os.walk(cwd):
        for name in file:
            if name.endswith('.csv'):
                add_csv(dir, name)
            elif name.endswith('.zip'):
                add_zip(dir, name)
                # print(ZipFile.namelist(os.path.join(dir, name)))
    if len(zip_files) > 0:
        unzipper(zip_files, csv_files)
    if len(csv_files) > 0:
        initialize_csv()
        # pass

if __name__ in '__main__':
    test_loop()
    # print(zip_files)
    # print(csv_files)
