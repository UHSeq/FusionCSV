# sort and merge multiple csv files to one single file for import into excel.

import os
from zipfile import ZipFile
import pandas
import openpyxl
from datetime import datetime

cwd = os.getcwd()
now = datetime.now()
nowtime = now.strftime('%H%M%S')
nowdate = now.strftime('%m%d%Y')
excel_file_name = f'fusions_{nowtime}_{nowdate}.xlsx'
output_dir = 'Output'
csv_hold = 'csv_hold'
excel_path = os.path.join(cwd, output_dir, excel_file_name)
count_col = 3
fusion_sheet = 'Total Fusions'

zip_files = []
csv_files = set()
csv_objs = {}


class CSVFiles():
    def __init__(self, path):
        self.path = path
        _, self.name = os.path.split(self.path)
        self.name, _ = os.path.splitext(self.name)
        self.read_data()
        self.find_groups()
        self.process_groups()
        # self.data.to_csv(os.path.join(cwd, csv_hold, self.name + '.csv'))
        # with pandas.ExcelWriter(excel_path, mode='a') as Excel_File:
        #     self.unique_counts.to_excel(Excel_File, self.name)


    def read_data(self):
        self.data = pandas.read_csv(self.path)

    def find_groups(self):
        self.groupby = self.data.groupby(['Name'])
        self.size = self.groupby.size()
        self.size = dict(self.size)
        self.data.insert(loc=count_col, column='Counts', value=0)

    
    def process_groups(self):
        for index, name in enumerate(self.data['Name']):
            self.data.iat[index, count_col] = self.size[name]
            # print(index, name)

# class Uniques():
#     def __init__(self, name) -> None:
#         self.name = name

class MasterData():
    def __init__(self, path):
        self.path = path
        self.create_excel()
        self.csv_files = {}
        self.fusions = {}
        self.csv_dict()
        self.fusion_count()
        self.print2excel()

    def csv_dict(self):
        for index, csv in enumerate(csv_files, 1):
            # print(f'####\nInitializing {index}\n{csv}\n####')
            self.csv_files[index] = CSVFiles(csv)
            # print()

    def create_excel(self):
        workbook = openpyxl.Workbook()
        workbook.save(self.path)
        # self.workbook = openpyxl.load_workbook(self.path)

    def remove_sheet(self):
        self.workbook = openpyxl.load_workbook(self.path)
        self.workbook.remove(self.workbook['Sheet'])
        self.workbook.save(self.path)

    def rename_sheet(self):
        self.wso = self.workbook['Sheet']
        self.wso.title = fusion_sheet
        self.workbook.save(self.path)
        # pass

    def fusion_count(self):
        for csv in self.csv_files.values():
            self.fusions[csv.name] = csv.size
        self.fusions = pandas.DataFrame.from_dict(self.fusions)
        # self.total_occurance = self.fusions.sum(axis=1)
        # self.fusions.loc[:,'Total Occurance'] = self.fusions.sum(axis=1)
        self.fusions.insert(0, 'Total Occurance', self.fusions.sum(axis=1))
        # self.fusions.to_csv(os.path.join(cwd, output_dir, 'temp_merged.csv'))

    def print2excel(self):
        # self.rename_sheet()
        # self.workbook = openpyxl.load_workbook(self.path)
        with pandas.ExcelWriter(self.path, mode='a') as Excel_File:
            # self.fusions.to_csv(os.path.join(cwd, output_dir, 'temp_merged.csv'))
            self.fusions.to_excel(Excel_File, fusion_sheet)
            for csv in self.csv_files.values():
                csv.data.to_excel(Excel_File, csv.name)
        self.remove_sheet()
        

def add_zip(directory, name, ziplist=zip_files):
    zip_path = write_pathname(directory, name)
    ziplist.append(zip_path)

def add_csv(directory, name, csvset=csv_files):
    csv_path = write_pathname(directory, name)
    csvset.add(csv_path)

def clear_folder(dir_path):
    pass
    for root, _, files in os.walk(dir_path):
        for file in files:
            os.remove(os.path.join(root, file))

# def initialize_csv():
#     # workbook = openpyxl.Workbook()
#     # workbook.save(excel_path)
#     for index, csv in enumerate(csv_files, 1):
#         print(f'#### Initializing {index} ####')
#         csv_objs[index] = CSVFiles(csv)
#     # workbook = openpyxl.load_workbook(excel_path)
#     # workbook.remove(workbook['Sheet'])
#     # workbook.save(excel_path)

def unzipper(zipperlist, csvset):
    for zipperfile in zipperlist:
        zipperpath = os.path.dirname(zipperfile)
        with ZipFile(zipperfile, 'r') as Zipper:
            tempset = set(ZipFile.namelist(Zipper))
            for file in tempset:
                csvset.add(os.path.join(zipperpath, file))
            Zipper.extractall(os.path.dirname(zipperfile))

def write_pathname(*strings):
    return os.path.join(*strings)

def main_loop():
    clear_folder(output_dir)
    clear_folder(csv_hold)
    for dir, _, file in os.walk(cwd):
        for name in file:
            if name.endswith('.csv'):
                add_csv(dir, name)
            elif name.endswith('.zip'):
                add_zip(dir, name)
    if len(zip_files) > 0:
        unzipper(zip_files, csv_files)
        # pass
    if len(csv_files) > 0:
        # initialize_csv()
        m = MasterData(excel_path)
        # pass
    else:
        print("####\nOpps... did you mean to add some csv files?\nNothing here...\n####")

if __name__ in '__main__':
    main_loop()
